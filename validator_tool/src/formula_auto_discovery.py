"""
Módulo para descoberta automática de fórmulas em arquivos Excel
"""

import logging
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import json


class FormulaAutoDiscovery:
    """Descobre automaticamente fórmulas em arquivos Excel"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self._workbook_cache = {}  # Cache para evitar recarregar arquivos
    
    def discover_formulas(self, file_path: str, 
                         sheet_name: Optional[str] = None,
                         sample_row: int = 2,
                         max_cols: int = 300) -> Dict[str, Dict[str, Any]]:
        """
        Descobre automaticamente todas as fórmulas em uma planilha
        
        Args:
            file_path: Caminho do arquivo Excel
            sheet_name: Nome da aba (None para aba ativa)
            sample_row: Linha de amostra para buscar fórmulas (padrão: 2)
            max_cols: Número máximo de colunas para verificar
            
        Returns:
            Dict com informações sobre cada fórmula encontrada
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        self.logger.info(f"Descobrindo fórmulas em: {file_path}")
        
        # Carregar workbook
        wb = load_workbook(file_path, data_only=False)
        
        # Selecionar planilha
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                raise ValueError(f"Aba '{sheet_name}' não encontrada")
        else:
            ws = wb.active
            sheet_name = ws.title
        
        formulas_found = {}
        
        # Verificar cada coluna na linha de amostra
        for col_idx in range(1, min(ws.max_column + 1, max_cols + 1)):
            # Obter header da coluna
            header = ws.cell(row=1, column=col_idx).value
            if not header:
                continue
            
            # Verificar se há fórmula na linha de amostra
            cell = ws.cell(row=sample_row, column=col_idx)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Analisar a fórmula
                formula_info = self._analyze_formula(formula, ws, col_letter)
                
                # Chave única para a fórmula
                formula_key = f"{sheet_name}.{header}" if sheet_name != "Sheet" else header
                
                formulas_found[formula_key] = {
                    'header': header,
                    'column': col_letter,
                    'formula_original': formula,
                    'formula_pattern': self._generalize_formula(formula),
                    'dependencies': formula_info['dependencies'],
                    'sheet_references': formula_info['sheet_references'],
                    'header_traducao': formula_info['header_traducao'],  # Mapeamento automático
                    'aba': sheet_name
                }
                
                self.logger.info(f"Fórmula encontrada em {col_letter}{sample_row} ({header}): {formula}")
        
        wb.close()
        
        return formulas_found
    
    def discover_formulas_lightweight(self, file_path: str,
                                    sheet_name: Optional[str] = None,
                                    sample_row: int = 2,
                                    columns_to_check: Optional[List[str]] = None,
                                    max_cols: int = 50) -> Dict[str, Dict[str, Any]]:
        """
        Versão otimizada para arquivos grandes - analisa apenas o necessário
        
        Args:
            file_path: Caminho do arquivo Excel
            sheet_name: Nome da aba (None para aba ativa)
            sample_row: Linha de amostra para buscar fórmulas
            columns_to_check: Lista específica de colunas para verificar (ex: ['A', 'B', 'F'])
            max_cols: Limite máximo de colunas se columns_to_check não for especificado
            
        Returns:
            Dict com informações sobre cada fórmula encontrada
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        self.logger.info(f"[LIGHTWEIGHT] Descobrindo fórmulas em: {file_path}")
        
        # Usar read_only=True para economizar memória
        wb = load_workbook(file_path, data_only=False, read_only=True)
        
        try:
            # Selecionar planilha
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    raise ValueError(f"Aba '{sheet_name}' não encontrada")
            else:
                ws = wb.active
                sheet_name = ws.title
            
            formulas_found = {}
            headers = {}  # Cache de headers
            
            # Se columns_to_check especificado, usar apenas essas colunas
            if columns_to_check:
                cols_to_process = []
                for col in columns_to_check:
                    if isinstance(col, str) and col.isalpha():
                        cols_to_process.append(openpyxl.utils.column_index_from_string(col))
                    elif isinstance(col, int):
                        cols_to_process.append(col)
            else:
                # Ler apenas primeira linha para determinar colunas com dados
                cols_to_process = []
                for col_idx in range(1, max_cols + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if header:
                        cols_to_process.append(col_idx)
                        headers[col_idx] = header
            
            self.logger.info(f"[LIGHTWEIGHT] Processando {len(cols_to_process)} colunas")
            
            # Processar apenas colunas selecionadas
            for col_idx in cols_to_process:
                # Obter header se ainda não temos
                if col_idx not in headers:
                    header = ws.cell(row=1, column=col_idx).value
                    if not header:
                        continue
                    headers[col_idx] = header
                else:
                    header = headers[col_idx]
                
                # Verificar fórmula na linha de amostra
                cell = ws.cell(row=sample_row, column=col_idx)
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # Análise simplificada para modo lightweight
                    formula_info = self._analyze_formula_lightweight(formula, headers)
                    
                    # Chave única para a fórmula
                    formula_key = f"{sheet_name}.{header}" if sheet_name != "Sheet" else header
                    
                    formulas_found[formula_key] = {
                        'header': header,
                        'column': col_letter,
                        'formula_original': formula,
                        'formula_pattern': self._generalize_formula(formula),
                        'dependencies': formula_info['dependencies'],
                        'sheet_references': formula_info['sheet_references'],
                        'header_traducao': formula_info['header_traducao'],
                        'aba': sheet_name
                    }
                    
                    self.logger.debug(f"[LIGHTWEIGHT] Fórmula em {col_letter}{sample_row}: {formula[:50]}...")
            
            self.logger.info(f"[LIGHTWEIGHT] {len(formulas_found)} fórmulas descobertas")
            return formulas_found
            
        finally:
            wb.close()
    
    def _analyze_formula_lightweight(self, formula: str, headers_cache: Dict[int, str]) -> Dict[str, Any]:
        """
        Análise simplificada de fórmula para modo lightweight
        Usa cache de headers para evitar acessos desnecessários ao worksheet
        """
        import re
        
        dependencies = []
        sheet_references = []
        header_traducao = {}
        
        # Padrões simplificados
        patterns = {
            'sheet_ref': re.compile(r'(\w+)!([A-Z]+)(\d+)'),
            'local_ref': re.compile(r'(?<![A-Z!])([A-Z]+)(\d+)(?![A-Z])'),
            'col_range': re.compile(r'([A-Z]+):[A-Z]+'),
            'sheet_col_range': re.compile(r'(\w+)!([A-Z]+):[A-Z]+')
        }
        
        # Processar referências com aba
        for match in patterns['sheet_ref'].finditer(formula):
            sheet = match.group(1)
            col = match.group(2)
            sheet_references.append(sheet)
            dependencies.append(f"{sheet}!{col}")
            header_traducao[f"{sheet}!{col}"] = f"{sheet}!{col}"  # Placeholder
        
        # Processar referências locais
        formula_clean = patterns['sheet_ref'].sub('', formula)
        for match in patterns['local_ref'].finditer(formula_clean):
            col = match.group(1)
            dependencies.append(col)
            # Tentar mapear usando cache se possível
            col_idx = openpyxl.utils.column_index_from_string(col)
            if col_idx in headers_cache:
                header_traducao[col] = headers_cache[col_idx]
            else:
                header_traducao[col] = col
        
        return {
            'dependencies': list(set(dependencies)),
            'sheet_references': list(set(sheet_references)),
            'header_traducao': header_traducao
        }
    
    def _analyze_formula(self, formula: str, worksheet, col_letter: str) -> Dict[str, Any]:
        """
        Analisa uma fórmula para extrair suas dependências e criar mapeamento de tradução
        """
        import re
        
        # Padrões para diferentes tipos de referências
        patterns = {
            'local_ref': re.compile(r'(?<![A-Z!])([A-Z]+)(\d+)(?![A-Z])'),  # Referências locais (A1, B2)
            'sheet_ref': re.compile(r'(\w+)!([A-Z]+)(\d+)'),      # Referências com aba (Sheet1!A1)
            'sheet_range': re.compile(r'(\w+)!([A-Z]+)(\d*):([A-Z]+)(\d*)'),  # Ranges com aba (Sheet1!A:A ou Sheet1!A1:B10)
            'sheet_col': re.compile(r'(\w+)!([A-Z]+):([A-Z]+)'),   # Colunas completas (Sheet1!A:A)
            'range_ref': re.compile(r'([A-Z]+\d+):([A-Z]+\d+)'),  # Ranges (A1:A10)
            'col_ref': re.compile(r'(?<![A-Z!:])([A-Z]+):([A-Z]+)(?![A-Z])'), # Colunas completas (A:A)
            'mixed_ref': re.compile(r'\$?([A-Z]+)\$?(\d+)')      # Referências mistas ($A$1, A$1, $A1)
        }
        
        dependencies = []
        sheet_references = []
        header_traducao = {}  # Mapeamento de tradução automático
        
        # Processar referências com colunas completas (Sheet!A:A)
        for match in patterns['sheet_col'].finditer(formula):
            sheet = match.group(1)
            col1 = match.group(2)
            col2 = match.group(3)
            
            sheet_references.append(sheet)
            
            # Para colunas completas, usar apenas a primeira coluna para o header
            try:
                if sheet in worksheet.parent.sheetnames:
                    ref_ws = worksheet.parent[sheet]
                    col_idx = openpyxl.utils.column_index_from_string(col1)
                    header = ref_ws.cell(row=1, column=col_idx).value
                    if header:
                        dependencies.append(f"{sheet}!{header}")
                        header_traducao[f"{sheet}!{col1}:{col2}"] = header
            except:
                pass
        
        # Processar ranges com aba (Sheet!A1:B10)
        for match in patterns['sheet_range'].finditer(formula):
            sheet = match.group(1)
            col1 = match.group(2)
            col2 = match.group(4)
            
            sheet_references.append(sheet)
            
            # Adicionar headers das colunas envolvidas
            try:
                if sheet in worksheet.parent.sheetnames:
                    ref_ws = worksheet.parent[sheet]
                    # Primeira coluna
                    col_idx1 = openpyxl.utils.column_index_from_string(col1)
                    header1 = ref_ws.cell(row=1, column=col_idx1).value
                    if header1:
                        dependencies.append(f"{sheet}!{header1}")
                    # Segunda coluna se diferente
                    if col1 != col2:
                        col_idx2 = openpyxl.utils.column_index_from_string(col2)
                        header2 = ref_ws.cell(row=1, column=col_idx2).value
                        if header2 and header2 != header1:
                            dependencies.append(f"{sheet}!{header2}")
            except:
                pass
        
        # Encontrar referências com aba simples
        for match in patterns['sheet_ref'].finditer(formula):
            sheet = match.group(1)
            col = match.group(2)
            row = match.group(3)
            
            # Verificar se já não foi processado como parte de um range
            if f"{sheet}!{col}{row}" in formula.replace(' ', ''):
                sheet_references.append(sheet)
                
                # Tentar obter o header da coluna referenciada
                try:
                    if sheet in worksheet.parent.sheetnames:
                        ref_ws = worksheet.parent[sheet]
                        col_idx = openpyxl.utils.column_index_from_string(col)
                        header = ref_ws.cell(row=1, column=col_idx).value
                        if header:
                            dependencies.append(f"{sheet}!{header}")
                            # Adicionar ao mapeamento de tradução
                            header_traducao[f"{sheet}!{col}"] = header
                        else:
                            dependencies.append(f"{sheet}!{col}")
                            header_traducao[f"{sheet}!{col}"] = f"{sheet}!{col}"
                    else:
                        dependencies.append(f"{sheet}!{col}")
                        header_traducao[f"{sheet}!{col}"] = f"{sheet}!{col}"
                except:
                    dependencies.append(f"{sheet}!{col}")
                    header_traducao[f"{sheet}!{col}"] = f"{sheet}!{col}"
        
        # Processar colunas completas locais (A:A)
        for match in patterns['col_ref'].finditer(formula):
            col1 = match.group(1)
            col2 = match.group(2)
            
            # Para colunas completas, usar apenas a primeira coluna
            try:
                col_idx = openpyxl.utils.column_index_from_string(col1)
                header = worksheet.cell(row=1, column=col_idx).value
                if header:
                    dependencies.append(header)
                    header_traducao[f"{col1}:{col2}"] = header
            except:
                pass
        
        # Encontrar referências locais
        # Limpar fórmula de elementos já processados
        formula_clean = formula
        for pattern_name in ['sheet_ref', 'sheet_range', 'sheet_col']:
            formula_clean = patterns[pattern_name].sub('', formula_clean)
        
        # Processar referências locais e mistas
        for match in patterns['mixed_ref'].finditer(formula_clean):
            col = match.group(1)
            
            # Ignorar se faz parte de uma função (ex: SUM, IF)
            pos = match.start()
            if pos > 0 and formula_clean[pos-1].isalpha():
                continue
            
            # Tentar obter o header
            try:
                col_idx = openpyxl.utils.column_index_from_string(col)
                header = worksheet.cell(row=1, column=col_idx).value
                if header:
                    dependencies.append(header)
                    # Adicionar ao mapeamento de tradução
                    header_traducao[col] = header
                    # Também mapear com $
                    header_traducao[f"${col}"] = header
                    header_traducao[f"{col}$"] = header
                    header_traducao[f"${col}$"] = header
                else:
                    dependencies.append(col)
                    header_traducao[col] = col
            except:
                dependencies.append(col)
                header_traducao[col] = col
        
        return {
            'dependencies': list(set(dependencies)),  # Remover duplicatas
            'sheet_references': list(set(sheet_references)),
            'header_traducao': header_traducao  # Mapeamento automático
        }
    
    def _generalize_formula(self, formula: str) -> str:
        """
        Generaliza uma fórmula substituindo números de linha por {row}
        Ex: =A2+B2 -> =A{row}+B{row}
        """
        import re
        
        # Substituir referências de linha por {row}
        pattern = re.compile(r'([A-Z]+)(\d+)')
        generalized = pattern.sub(r'\1{row}', formula)
        
        return generalized
    
    def discover_all_sheets(self, file_path: str, 
                           sample_row: int = 2) -> Dict[str, Dict[str, Any]]:
        """
        Descobre fórmulas em todas as abas de um arquivo
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        wb = load_workbook(file_path, data_only=False)
        all_formulas = {}
        
        for sheet_name in wb.sheetnames:
            self.logger.info(f"Analisando aba: {sheet_name}")
            
            try:
                sheet_formulas = self.discover_formulas(
                    file_path, 
                    sheet_name=sheet_name,
                    sample_row=sample_row
                )
                all_formulas.update(sheet_formulas)
            except Exception as e:
                self.logger.warning(f"Erro ao analisar aba {sheet_name}: {e}")
        
        wb.close()
        
        return all_formulas
    
    def discover_all_sheets_lightweight(self, file_path: str,
                                      sample_row: int = 2,
                                      columns_per_sheet: Optional[Dict[str, List[str]]] = None,
                                      max_cols_per_sheet: int = 50) -> Dict[str, Dict[str, Any]]:
        """
        Versão otimizada para descobrir fórmulas em todas as abas de arquivos grandes
        
        Args:
            file_path: Caminho do arquivo Excel
            sample_row: Linha de amostra para buscar fórmulas
            columns_per_sheet: Dict especificando colunas por aba ex: {'Sheet1': ['A', 'B', 'F']}
            max_cols_per_sheet: Limite de colunas por aba se não especificado
            
        Returns:
            Dict com todas as fórmulas descobertas
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        self.logger.info(f"[LIGHTWEIGHT] Descobrindo fórmulas em todas as abas de: {file_path}")
        
        # Usar read_only para economizar memória
        wb = load_workbook(file_path, data_only=False, read_only=True)
        all_formulas = {}
        
        try:
            for sheet_name in wb.sheetnames:
                self.logger.info(f"[LIGHTWEIGHT] Analisando aba: {sheet_name}")
                
                # Determinar colunas para esta aba
                cols_to_check = None
                if columns_per_sheet and sheet_name in columns_per_sheet:
                    cols_to_check = columns_per_sheet[sheet_name]
                
                try:
                    sheet_formulas = self.discover_formulas_lightweight(
                        file_path,
                        sheet_name=sheet_name,
                        sample_row=sample_row,
                        columns_to_check=cols_to_check,
                        max_cols=max_cols_per_sheet
                    )
                    all_formulas.update(sheet_formulas)
                except Exception as e:
                    self.logger.warning(f"[LIGHTWEIGHT] Erro ao analisar aba {sheet_name}: {e}")
            
            self.logger.info(f"[LIGHTWEIGHT] Total de fórmulas descobertas: {len(all_formulas)}")
            return all_formulas
            
        finally:
            wb.close()
    
    def generate_config_section(self, formulas: Dict[str, Dict[str, Any]], 
                               file_path: str) -> Dict[str, Any]:
        """
        Gera uma seção de configuração baseada nas fórmulas descobertas
        """
        config_section = {
            "caminho": file_path,
            "tipo": "auto_discovery",
            "formulas_descobertas": {}
        }
        
        # Agrupar por aba
        formulas_by_sheet = {}
        
        for formula_key, formula_info in formulas.items():
            sheet = formula_info['aba']
            if sheet not in formulas_by_sheet:
                formulas_by_sheet[sheet] = {
                    "aba_planilha": sheet,
                    "formulas": []
                }
            
            # Criar mapeamento automático
            formula_config = {
                "header_resultado": formula_info['header'],
                "coluna": formula_info['column'],
                "formula_descoberta": formula_info['formula_original'],
                "formula_pattern": formula_info['formula_pattern'],
                "dependencies": formula_info['dependencies'],
                "header_traducao": formula_info.get('header_traducao', {})
            }
            
            formulas_by_sheet[sheet]["formulas"].append(formula_config)
        
        config_section["abas"] = formulas_by_sheet
        
        return config_section
    
    def save_discovered_config(self, formulas: Dict[str, Dict[str, Any]], 
                              output_path: str = "formulas_descobertas.json") -> None:
        """
        Salva as fórmulas descobertas em um arquivo JSON
        """
        config = {
            "descricao": "Fórmulas descobertas automaticamente",
            "formulas": formulas,
            "total_formulas": len(formulas)
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        self.logger.info(f"Configuração salva em: {output_path}")