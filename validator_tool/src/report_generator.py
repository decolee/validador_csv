"""
Módulo para geração de relatórios XLSX com formatação
"""

import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """Gera relatórios XLSX formatados com resultados da validação"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Estilos padrão
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self, validation_results: List[Dict[str, Any]], 
                formulas: Dict[str, Dict[str, Any]],
                unmatched: Dict[str, List],
                output_path: str,
                extra_info: Optional[Dict[str, Any]] = None) -> None:
        """
        Gera relatório completo de validação
        
        Args:
            validation_results: Resultados da validação
            formulas: Fórmulas extraídas e traduzidas
            unmatched: Linhas sem correspondência
            output_path: Caminho para salvar o relatório
        """
        # Criar diretório de saída se não existir
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Criar workbook
        wb = Workbook()
        
        # Remover planilha padrão
        wb.remove(wb.active)
        
        # 1. Criar aba de sumário
        self._create_summary_sheet(wb, validation_results, unmatched)
        
        # 2. Criar aba de detalhes da validação
        self._create_validation_details_sheet(wb, validation_results, formulas)
        
        # 3. Criar aba de linhas sem correspondência (se houver)
        if unmatched['df1'] or unmatched['df2']:
            self._create_unmatched_sheet(wb, unmatched)
        
        # 4. Criar aba de fórmulas (se houver)
        if formulas:
            self._create_formulas_sheet(wb, formulas)
        
        # 5. Criar abas extras com informações úteis
        if extra_info:
            # Mapa de dependências
            if 'dependencias' in extra_info:
                self._create_dependency_map_sheet(wb, extra_info['dependencias'])
            
            # Alertas e recomendações
            if 'alertas' in extra_info:
                self._create_alerts_sheet(wb, extra_info['alertas'])
            
            # Análise de impacto em cascata
            if 'impacto_cascata' in extra_info:
                self._create_cascade_impact_sheet(wb, extra_info['impacto_cascata'])
            
            # Resumo executivo
            self._create_executive_summary(wb, validation_results, formulas, extra_info)
        
        # Salvar arquivo
        wb.save(output_path)
        self.logger.info(f"Relatório salvo em: {output_path}")
    
    def _create_summary_sheet(self, wb: Workbook, validation_results: List[Dict[str, Any]], 
                            unmatched: Dict[str, List]) -> None:
        """Cria aba de sumário com estatísticas gerais"""
        ws = wb.create_sheet("Sumário")
        
        # Converter resultados para DataFrame para facilitar análise
        df_results = pd.DataFrame(validation_results) if validation_results else pd.DataFrame()
        
        # Calcular estatísticas
        total_comparacoes = len(validation_results)
        total_divergencias = len(df_results[df_results['resultado'] == False]) if not df_results.empty else 0
        total_concordancias = total_comparacoes - total_divergencias
        taxa_concordancia = (total_concordancias / total_comparacoes * 100) if total_comparacoes > 0 else 0
        
        # Dados do sumário
        summary_data = [
            ["RELATÓRIO DE VALIDAÇÃO DE DADOS", ""],
            ["", ""],
            ["Estatísticas Gerais", ""],
            ["Total de Comparações", total_comparacoes],
            ["Total de Concordâncias", total_concordancias],
            ["Total de Divergências", total_divergencias],
            ["Taxa de Concordância", f"{taxa_concordancia:.2f}%"],
            ["", ""],
            ["Linhas sem Correspondência", ""],
            ["Apenas no Arquivo 1", len(unmatched.get('df1', []))],
            ["Apenas no Arquivo 2", len(unmatched.get('df2', []))]
        ]
        
        # Adicionar dados à planilha
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Aplicar formatação depois
        # Formatar título
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells(f"A1:B1")
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Formatar cabeçalhos de seção
        for row_idx, row_data in enumerate(summary_data, 1):
            if row_data[0] in ["Estatísticas Gerais", "Linhas sem Correspondência"]:
                ws.cell(row=row_idx, column=1).font = Font(bold=True, color="366092")
                ws.merge_cells(f"A{row_idx}:B{row_idx}")
        
        # Adicionar divergências por coluna
        if not df_results.empty:
            ws.cell(row=13, column=1, value="Divergências por Coluna")
            ws.cell(row=13, column=1).font = Font(bold=True, color="366092")
            ws.merge_cells("A13:B13")
            
            row_idx = 14
            for col in df_results['coluna'].unique():
                col_data = df_results[df_results['coluna'] == col]
                divergencias = len(col_data[col_data['resultado'] == False])
                ws.cell(row=row_idx, column=1, value=col)
                ws.cell(row=row_idx, column=2, value=divergencias)
                row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_validation_details_sheet(self, wb: Workbook, validation_results: List[Dict[str, Any]],
                                       formulas: Dict[str, Dict[str, Any]]) -> None:
        """Cria aba com detalhes da validação"""
        ws = wb.create_sheet("Detalhes da Validação")
        
        if not validation_results:
            ws.cell(row=1, column=1, value="Nenhum resultado de validação disponível")
            return
        
        # Preparar dados para o relatório
        report_data = []
        
        # Agrupar por linha para facilitar visualização
        df_results = pd.DataFrame(validation_results)
        
        for _, row in df_results.iterrows():
            # Obter fórmula se disponível
            coluna_nome = row['coluna']
            
            # Tentar buscar fórmula com diferentes formatos de chave
            formula_info = formulas.get(coluna_nome, {})
            
            # Se não encontrou, tentar com prefixo de aba
            if not formula_info:
                # Procurar por qualquer chave que termine com o nome da coluna
                for key in formulas.keys():
                    if key.endswith(f'.{coluna_nome}') or key == coluna_nome:
                        formula_info = formulas[key]
                        break
            
            # Debug
            if formula_info:
                self.logger.debug(f"Fórmula encontrada para coluna '{coluna_nome}': {formula_info.get('formula_original', 'N/A')}")
            
            # Preparar fórmulas como texto (adicionar apóstrofo para evitar cálculo)
            formula_original = formula_info.get('formula_original', '')
            formula_traduzida = formula_info.get('formula_traduzida', '')
            
            # Se houver fórmula, adicionar apóstrofo no início para forçar como texto
            if formula_original and formula_original.startswith('='):
                formula_original = "'" + formula_original
            if formula_traduzida and formula_traduzida.startswith('='):
                formula_traduzida = "'" + formula_traduzida
            
            report_row = {
                'Chave': row.get('chave', ''),
                'Linha': row.get('linha_idx', ''),
                'Coluna_Validada': row['coluna'],
                'Valor_Arquivo_1': row['valor_arquivo_1'],
                'Valor_Arquivo_2': row['valor_arquivo_2'],
                'Resultado_Validacao': 'VERDADEIRO' if row['resultado'] else 'FALSO',
                'Tolerancia_Aplicada': row['tolerancia_aplicada'],
                'Formula_Original': formula_original,
                'Formula_Traduzida': formula_traduzida,
                'Divergencias_Na_Formula': row.get('divergencias_na_formula') if row.get('divergencias_na_formula') else ''
            }
            
            report_data.append(report_row)
        
        # Converter para DataFrame e adicionar à planilha
        df_report = pd.DataFrame(report_data)
        
        # Adicionar headers
        headers = list(df_report.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border
        
        # Adicionar dados
        for row_idx, row_data in enumerate(dataframe_to_rows(df_report, index=False, header=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = Alignment(vertical="center")
                
                # Aplicar formatação condicional para coluna de resultado
                if col_idx == 6:  # Coluna Resultado_Validacao
                    if value == 'VERDADEIRO':
                        cell.fill = self.success_fill
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = self.error_fill
                        cell.font = Font(color="9C0006")
                
                # Destacar divergências na fórmula
                if col_idx == 10 and value:  # Coluna Divergencias_Na_Formula
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(color="CC6600", bold=True)
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_unmatched_sheet(self, wb: Workbook, unmatched: Dict[str, List]) -> None:
        """Cria aba com linhas sem correspondência"""
        ws = wb.create_sheet("Linhas sem Correspondência")
        
        # Headers
        headers = ["Valor da Chave", "Presente em", "Observação"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row_idx = 2
        
        # Adicionar linhas apenas no arquivo 1
        for key in unmatched.get('df1', []):
            ws.cell(row=row_idx, column=1, value=key).border = self.border
            ws.cell(row=row_idx, column=2, value="Apenas Arquivo 1").border = self.border
            ws.cell(row=row_idx, column=3, value="Sem correspondência no Arquivo 2").border = self.border
            row_idx += 1
        
        # Adicionar linhas apenas no arquivo 2
        for key in unmatched.get('df2', []):
            ws.cell(row=row_idx, column=1, value=key).border = self.border
            ws.cell(row=row_idx, column=2, value="Apenas Arquivo 2").border = self.border
            ws.cell(row=row_idx, column=3, value="Sem correspondência no Arquivo 1").border = self.border
            row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_formulas_sheet(self, wb: Workbook, formulas: Dict[str, Dict[str, Any]]) -> None:
        """Cria aba com detalhes das fórmulas extraídas"""
        ws = wb.create_sheet("Fórmulas Extraídas")
        
        # Headers
        headers = ["Coluna", "Fórmula Original", "Fórmula Traduzida", "Mapeamento Usado"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        # Adicionar dados das fórmulas
        row_idx = 2
        for col_name, formula_info in formulas.items():
            # Preparar fórmulas como texto
            formula_orig = formula_info.get('formula_original', '')
            formula_trad = formula_info.get('formula_traduzida', '')
            
            # Adicionar apóstrofo para forçar como texto
            if formula_orig.startswith('='):
                formula_orig = "'" + formula_orig
            if formula_trad.startswith('='):
                formula_trad = "'" + formula_trad
            
            ws.cell(row=row_idx, column=1, value=col_name).border = self.border
            ws.cell(row=row_idx, column=2, value=formula_orig).border = self.border
            ws.cell(row=row_idx, column=3, value=formula_trad).border = self.border
            
            # Converter mapeamento para string legível
            mapping_str = ", ".join([f"{k}={v}" for k, v in formula_info.get('mapeamento', {}).items()])
            ws.cell(row=row_idx, column=4, value=mapping_str).border = self.border
            
            row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_dependency_map_sheet(self, wb: Workbook, dependencies: Dict[str, Any]) -> None:
        """Cria aba com mapa visual de dependências entre colunas e fórmulas"""
        ws = wb.create_sheet("Mapa de Dependências")
        
        # Headers
        headers = ["Coluna/Fórmula", "Depende De", "É Usada Por", "Tipo Dependência", "Impacto"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row_idx = 2
        
        # Processar dependências de fórmulas
        if 'formulas' in dependencies:
            for col_name, deps in dependencies['formulas'].items():
                # Dependências diretas
                if 'depends_on' in deps:
                    for dep in deps['depends_on']:
                        ws.cell(row=row_idx, column=1, value=col_name).border = self.border
                        ws.cell(row=row_idx, column=2, value=dep).border = self.border
                        ws.cell(row=row_idx, column=3, value="-").border = self.border
                        ws.cell(row=row_idx, column=4, value="Fórmula → Coluna").border = self.border
                        ws.cell(row=row_idx, column=5, value="Direto").border = self.border
                        row_idx += 1
                
                # Referências externas (outras abas)
                if 'referencias_externas' in deps:
                    for ref_externa in deps['referencias_externas']:
                        ws.cell(row=row_idx, column=1, value=col_name).border = self.border
                        ws.cell(row=row_idx, column=2, value=f"{ref_externa}!*").border = self.border
                        ws.cell(row=row_idx, column=3, value="-").border = self.border
                        ws.cell(row=row_idx, column=4, value="Cross-Sheet").border = self.border
                        ws.cell(row=row_idx, column=5, value="Externa").border = self.border
                        row_idx += 1
        
        # Processar impactos reversos
        if 'reverse_dependencies' in dependencies:
            for col_name, impacted_by in dependencies['reverse_dependencies'].items():
                for formula in impacted_by:
                    ws.cell(row=row_idx, column=1, value=col_name).border = self.border
                    ws.cell(row=row_idx, column=2, value="-").border = self.border
                    ws.cell(row=row_idx, column=3, value=formula).border = self.border
                    ws.cell(row=row_idx, column=4, value="Coluna → Fórmula").border = self.border
                    ws.cell(row=row_idx, column=5, value="Reverso").border = self.border
                    row_idx += 1
        
        # Adicionar resumo de estatísticas
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="ESTATÍSTICAS DE DEPENDÊNCIAS")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="366092")
        ws.merge_cells(f"A{row_idx}:E{row_idx}")
        
        row_idx += 1
        stats = dependencies.get('statistics', {})
        stats_data = [
            ["Total de Fórmulas", stats.get('total_formulas', 0)],
            ["Total de Dependências", stats.get('total_dependencies', 0)],
            ["Dependências Cross-Sheet", stats.get('cross_sheet_deps', 0)],
            ["Colunas mais referenciadas", ", ".join(stats.get('most_referenced', []))],
            ["Profundidade máxima de dependência", stats.get('max_depth', 0)]
        ]
        
        for stat in stats_data:
            ws.cell(row=row_idx, column=1, value=stat[0]).border = self.border
            ws.cell(row=row_idx, column=2, value=stat[1]).border = self.border
            ws.merge_cells(f"B{row_idx}:E{row_idx}")
            row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_alerts_sheet(self, wb: Workbook, alerts: List[Dict[str, Any]]) -> None:
        """Cria aba com alertas e recomendações baseados na análise"""
        ws = wb.create_sheet("Alertas e Recomendações")
        
        # Headers
        headers = ["Severidade", "Tipo", "Localização", "Descrição", "Recomendação", "Impacto Estimado"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        if not alerts:
            ws.cell(row=2, column=1, value="Nenhum alerta detectado")
            ws.merge_cells("A2:F2")
            return
        
        # Ordenar alertas por severidade
        severity_order = {'alta': 1, 'media': 2, 'baixa': 3}
        sorted_alerts = sorted(alerts, key=lambda x: severity_order.get(x.get('severidade', 'baixa'), 4))
        
        row_idx = 2
        for alert in sorted_alerts:
            # Severidade com formatação condicional
            sev_cell = ws.cell(row=row_idx, column=1, value=alert.get('severidade', '').upper())
            sev_cell.border = self.border
            sev_cell.alignment = Alignment(horizontal="center")
            
            if alert.get('severidade') == 'alta':
                sev_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                sev_cell.font = Font(color="FFFFFF", bold=True)
            elif alert.get('severidade') == 'media':
                sev_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                sev_cell.font = Font(color="FFFFFF", bold=True)
            else:
                sev_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                sev_cell.font = Font(bold=True)
            
            # Outros campos
            ws.cell(row=row_idx, column=2, value=alert.get('tipo', '')).border = self.border
            ws.cell(row=row_idx, column=3, value=alert.get('localizacao', '')).border = self.border
            ws.cell(row=row_idx, column=4, value=alert.get('descricao', '')).border = self.border
            ws.cell(row=row_idx, column=5, value=alert.get('recomendacao', '')).border = self.border
            ws.cell(row=row_idx, column=6, value=alert.get('impacto', '')).border = self.border
            
            row_idx += 1
        
        # Adicionar resumo de alertas
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="RESUMO DE ALERTAS")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        
        row_idx += 1
        # Contar alertas por severidade
        alta_count = len([a for a in alerts if a.get('severidade') == 'alta'])
        media_count = len([a for a in alerts if a.get('severidade') == 'media'])
        baixa_count = len([a for a in alerts if a.get('severidade') == 'baixa'])
        
        summary_data = [
            ["Alertas de Alta Severidade", alta_count],
            ["Alertas de Média Severidade", media_count],
            ["Alertas de Baixa Severidade", baixa_count],
            ["Total de Alertas", len(alerts)]
        ]
        
        for summary_row in summary_data:
            ws.cell(row=row_idx, column=1, value=summary_row[0]).border = self.border
            ws.cell(row=row_idx, column=2, value=summary_row[1]).border = self.border
            ws.merge_cells(f"C{row_idx}:F{row_idx}")
            row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_cascade_impact_sheet(self, wb: Workbook, cascade_data: Dict[str, Any]) -> None:
        """Cria aba mostrando o impacto em cascata de divergências"""
        ws = wb.create_sheet("Impacto em Cascata")
        
        # Headers
        headers = ["Coluna Origem", "Valor Divergente", "Colunas Impactadas", "Fórmulas Afetadas", 
                   "Nível de Impacto", "Estimativa de Correção"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        row_idx = 2
        
        # Processar dados de impacto em cascata
        for origem, impactos in cascade_data.items():
            if isinstance(impactos, dict) and 'affected' in impactos:
                # Coluna origem
                ws.cell(row=row_idx, column=1, value=origem).border = self.border
                
                # Valor divergente (se disponível)
                ws.cell(row=row_idx, column=2, value=impactos.get('divergent_value', 'N/A')).border = self.border
                
                # Colunas impactadas
                affected_cols = ", ".join(impactos['affected'].get('columns', []))
                ws.cell(row=row_idx, column=3, value=affected_cols).border = self.border
                
                # Fórmulas afetadas
                affected_formulas = ", ".join(impactos['affected'].get('formulas', []))
                ws.cell(row=row_idx, column=4, value=affected_formulas).border = self.border
                
                # Nível de impacto
                impact_level = len(impactos['affected'].get('columns', [])) + len(impactos['affected'].get('formulas', []))
                if impact_level > 5:
                    nivel = "CRÍTICO"
                    cell_nivel = ws.cell(row=row_idx, column=5, value=nivel)
                    cell_nivel.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell_nivel.font = Font(color="FFFFFF", bold=True)
                elif impact_level > 2:
                    nivel = "ALTO"
                    cell_nivel = ws.cell(row=row_idx, column=5, value=nivel)
                    cell_nivel.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    cell_nivel.font = Font(color="FFFFFF")
                else:
                    nivel = "MODERADO"
                    cell_nivel = ws.cell(row=row_idx, column=5, value=nivel)
                    cell_nivel.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                cell_nivel.border = self.border
                cell_nivel.alignment = Alignment(horizontal="center")
                
                # Estimativa de correção
                if impact_level > 5:
                    estimativa = "Requer revisão completa do processo"
                elif impact_level > 2:
                    estimativa = "Verificar origem e recalcular dependências"
                else:
                    estimativa = "Correção pontual necessária"
                ws.cell(row=row_idx, column=6, value=estimativa).border = self.border
                
                row_idx += 1
        
        # Adicionar análise de propagação
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="ANÁLISE DE PROPAGAÇÃO DE ERROS")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        
        row_idx += 1
        
        # Estatísticas de propagação
        total_origins = len(cascade_data)
        total_affected = sum(len(v.get('affected', {}).get('columns', [])) + 
                           len(v.get('affected', {}).get('formulas', [])) 
                           for v in cascade_data.values() if isinstance(v, dict))
        
        propagation_stats = [
            ["Total de Colunas com Divergências", total_origins],
            ["Total de Colunas/Fórmulas Impactadas", total_affected],
            ["Taxa de Propagação Média", f"{total_affected/total_origins:.1f}x" if total_origins > 0 else "0x"],
            ["Colunas Críticas (>5 impactos)", len([k for k, v in cascade_data.items() 
                                                   if isinstance(v, dict) and 
                                                   len(v.get('affected', {}).get('columns', [])) + 
                                                   len(v.get('affected', {}).get('formulas', [])) > 5])]
        ]
        
        for stat in propagation_stats:
            ws.cell(row=row_idx, column=1, value=stat[0]).border = self.border
            ws.cell(row=row_idx, column=2, value=stat[1]).border = self.border
            ws.merge_cells(f"C{row_idx}:F{row_idx}")
            row_idx += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 35
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
    
    def _create_executive_summary(self, wb: Workbook, validation_results: List[Dict[str, Any]], 
                                formulas: Dict[str, Dict[str, Any]], extra_info: Dict[str, Any]) -> None:
        """Cria resumo executivo com principais insights e recomendações"""
        ws = wb.create_sheet("Resumo Executivo", 0)  # Inserir como primeira aba
        
        # Título
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value="RESUMO EXECUTIVO - VALIDAÇÃO DE DADOS")
        title_cell.font = Font(size=16, bold=True, color="366092")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data e hora
        from datetime import datetime
        ws.merge_cells("A2:F2")
        date_cell = ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        date_cell.alignment = Alignment(horizontal="center")
        
        row_idx = 4
        
        # 1. Visão Geral
        ws.cell(row=row_idx, column=1, value="1. VISÃO GERAL")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        row_idx += 1
        
        # Estatísticas principais
        df_results = pd.DataFrame(validation_results) if validation_results else pd.DataFrame()
        total_validacoes = len(validation_results)
        total_divergencias = len(df_results[df_results['resultado'] == False]) if not df_results.empty else 0
        taxa_sucesso = ((total_validacoes - total_divergencias) / total_validacoes * 100) if total_validacoes > 0 else 0
        
        overview_data = [
            ["Total de Validações Realizadas:", total_validacoes],
            ["Taxa de Sucesso:", f"{taxa_sucesso:.1f}%"],
            ["Total de Divergências:", total_divergencias],
            ["Colunas com Fórmulas:", len(formulas)],
            ["Tipo de Validação:", extra_info.get('tipo_validacao', 'Padrão')]
        ]
        
        for item in overview_data:
            ws.cell(row=row_idx, column=1, value=item[0]).border = self.border
            ws.cell(row=row_idx, column=2, value=item[1]).border = self.border
            ws.merge_cells(f"C{row_idx}:F{row_idx}")
            row_idx += 1
        
        row_idx += 1
        
        # 2. Principais Problemas Encontrados
        ws.cell(row=row_idx, column=1, value="2. PRINCIPAIS PROBLEMAS ENCONTRADOS")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        row_idx += 1
        
        if not df_results.empty and total_divergencias > 0:
            # Top 5 colunas com mais divergências
            col_errors = df_results[df_results['resultado'] == False]['coluna'].value_counts().head(5)
            
            for col, count in col_errors.items():
                pct = (count / total_validacoes * 100)
                ws.cell(row=row_idx, column=1, value=f"• {col}:").border = self.border
                ws.cell(row=row_idx, column=2, value=f"{count} divergências ({pct:.1f}%)").border = self.border
                ws.merge_cells(f"C{row_idx}:F{row_idx}")
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value="• Nenhuma divergência encontrada")
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            row_idx += 1
        
        row_idx += 1
        
        # 3. Análise de Impacto
        ws.cell(row=row_idx, column=1, value="3. ANÁLISE DE IMPACTO")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        row_idx += 1
        
        # Alertas críticos
        if 'alertas' in extra_info:
            alertas_criticos = [a for a in extra_info['alertas'] if a.get('severidade') == 'alta']
            if alertas_criticos:
                for alert in alertas_criticos[:3]:  # Top 3 alertas críticos
                    ws.cell(row=row_idx, column=1, value=f"• CRÍTICO: {alert.get('descricao', '')}").border = self.border
                    ws.merge_cells(f"A{row_idx}:F{row_idx}")
                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value="• Nenhum alerta crítico identificado")
                ws.merge_cells(f"A{row_idx}:F{row_idx}")
                row_idx += 1
        
        row_idx += 1
        
        # 4. Recomendações
        ws.cell(row=row_idx, column=1, value="4. RECOMENDAÇÕES")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        row_idx += 1
        
        # Gerar recomendações baseadas nos resultados
        if taxa_sucesso < 90:
            ws.cell(row=row_idx, column=1, value="• Revisar processo de migração - taxa de sucesso abaixo de 90%")
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            row_idx += 1
        
        if total_divergencias > 0 and not df_results.empty:
            # Verificar se há padrão nas divergências
            top_error_col = df_results[df_results['resultado'] == False]['coluna'].value_counts().index[0]
            ws.cell(row=row_idx, column=1, value=f"• Priorizar correção da coluna '{top_error_col}' - maior número de divergências")
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            row_idx += 1
        
        if 'impacto_cascata' in extra_info and len(extra_info['impacto_cascata']) > 0:
            ws.cell(row=row_idx, column=1, value="• Verificar dependências entre colunas - erros estão se propagando")
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            row_idx += 1
        
        # 5. Próximos Passos
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="5. PRÓXIMOS PASSOS")
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        row_idx += 1
        
        next_steps = [
            "1. Revisar detalhes das divergências na aba 'Detalhes da Validação'",
            "2. Analisar impacto em cascata na aba 'Impacto em Cascata'",
            "3. Verificar alertas e recomendações específicas",
            "4. Corrigir divergências começando pelas colunas de maior impacto",
            "5. Reexecutar validação após correções"
        ]
        
        for step in next_steps:
            ws.cell(row=row_idx, column=1, value=step)
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            row_idx += 1
        
        # Formatar células
        for row in ws.iter_rows(min_row=4, max_row=row_idx-1, min_col=1, max_col=6):
            for cell in row:
                if cell.value and not cell.border.left.style:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
        
        # Adicionar bordas ao redor do resumo
        thick_border = Border(
            left=Side(style='thick', color="366092"),
            right=Side(style='thick', color="366092"),
            top=Side(style='thick', color="366092"),
            bottom=Side(style='thick', color="366092")
        )
        
        # Aplicar borda externa
        for row in range(1, row_idx):
            ws.cell(row=row, column=1).border = Border(left=Side(style='thick', color="366092"))
            ws.cell(row=row, column=6).border = Border(right=Side(style='thick', color="366092"))
        
        for col in range(1, 7):
            ws.cell(row=1, column=col).border = Border(top=Side(style='thick', color="366092"))
            ws.cell(row=row_idx-1, column=col).border = Border(bottom=Side(style='thick', color="366092"))