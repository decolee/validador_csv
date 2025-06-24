"""
Módulo para extração e tradução de fórmulas Excel
"""

import logging
import re
from typing import Dict, Any, List, Tuple, Optional
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
import pandas as pd
from .formula_auto_discovery import FormulaAutoDiscovery


class FormulaExtractor:
    """Extrai e traduz fórmulas de arquivos Excel"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        # Padrão regex para identificar referências de célula
        self.cell_ref_pattern = re.compile(r'\b([A-Z]+)(\d+)\b')
    
    def extract_and_translate(self, config: Dict[str, Any], key_column: str) -> Dict[str, Dict[str, Any]]:
        """
        Extrai e traduz fórmulas conforme configuração
        Suporta configuração simples, multi-abas e descoberta automática
        """
        # Verificar tipo de configuração
        tipo = config.get('tipo')
        
        if tipo == 'auto_discovery':
            return self._extract_auto_discovery(config, key_column)
        elif tipo == 'multi_abas':
            return self._extract_multi_sheet(config, key_column)
        else:
            return self._extract_single_sheet(config, key_column)
    
    def _extract_single_sheet(self, config: Dict[str, Any], key_column: str) -> Dict[str, Dict[str, Any]]:
        """
        Extrai e traduz fórmulas conforme configuração
        
        Args:
            config: Configuração do arquivo de fórmulas
            key_column: Nome da coluna chave para indexação
            
        Returns:
            Dicionário com fórmulas extraídas e traduzidas
        """
        file_path = config['caminho']
        sheet_name = config.get('aba_planilha')
        columns_to_load = config.get('colunas_para_carregar', [])
        mappings = config.get('mapeamento_formulas', [])
        
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo de fórmulas não encontrado: {file_path}")
        
        self.logger.info(f"Abrindo arquivo de fórmulas: {file_path}")
        
        # Carregar workbook com openpyxl para acessar fórmulas
        wb = load_workbook(file_path, data_only=False)
        
        # Selecionar planilha
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                raise ValueError(f"Aba '{sheet_name}' não encontrada no arquivo")
        else:
            ws = wb.active
        
        # Extrair fórmulas para cada mapeamento
        formulas_dict = {}
        
        for mapping in mappings:
            header_resultado = mapping['header_resultado']
            header_traducao = mapping['header_traducao']
            
            self.logger.info(f"Processando fórmulas para coluna: {header_resultado}")
            
            # Encontrar coluna com o header
            col_idx = self._find_column_by_header(ws, header_resultado)
            if col_idx is None:
                self.logger.warning(f"Header '{header_resultado}' não encontrado")
                continue
            
            # Extrair fórmula da linha 2 (assumindo que linha 1 tem headers)
            cell = ws.cell(row=2, column=col_idx)
            
            # Em openpyxl, fórmulas são armazenadas no atributo 'value' quando data_only=False
            formula_original = None
            if hasattr(cell, 'value') and cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_original = cell.value
                elif hasattr(cell, '_value') and isinstance(cell._value, str) and str(cell._value).startswith('='):
                    formula_original = cell._value
            
            if formula_original:
                self.logger.debug(f"Fórmula encontrada: {formula_original}")
                
                # Traduzir fórmula
                formula_traduzida = self._translate_formula(formula_original, header_traducao)
                
                # Armazenar resultado
                formulas_dict[header_resultado] = {
                    'formula_original': formula_original,
                    'formula_traduzida': formula_traduzida,
                    'coluna_idx': col_idx,
                    'mapeamento': header_traducao
                }
                self.logger.info(f"Fórmula extraída para {header_resultado}: {formula_original}")
            else:
                self.logger.warning(f"Nenhuma fórmula encontrada na célula {openpyxl.utils.get_column_letter(col_idx)}2 para coluna '{header_resultado}'")
        
        wb.close()
        
        return formulas_dict
    
    def _extract_multi_sheet(self, config: Dict[str, Any], key_column: str) -> Dict[str, Dict[str, Any]]:
        """
        Extrai fórmulas de configuração multi-abas
        """
        file_path = config['caminho']
        abas_config = config.get('abas', {})
        
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo de fórmulas não encontrado: {file_path}")
        
        self.logger.info(f"Abrindo arquivo multi-abas: {file_path}")
        
        all_formulas = {}
        
        # Carregar workbook
        wb = load_workbook(file_path, data_only=False)
        
        # Processar cada aba configurada
        for aba_nome, aba_config in abas_config.items():
            if 'mapeamento_formulas' not in aba_config:
                continue
                
            self.logger.info(f"Processando aba: {aba_nome}")
            
            # Selecionar aba
            if aba_nome in wb.sheetnames:
                ws = wb[aba_nome]
            else:
                self.logger.warning(f"Aba '{aba_nome}' não encontrada")
                continue
            
            # Processar mapeamentos desta aba
            for mapping in aba_config['mapeamento_formulas']:
                header_resultado = mapping['header_resultado']
                header_traducao = mapping['header_traducao']
                
                # Encontrar coluna
                col_idx = self._find_column_by_header(ws, header_resultado)
                if col_idx is None:
                    self.logger.warning(f"Header '{header_resultado}' não encontrado na aba {aba_nome}")
                    continue
                
                # Extrair fórmula
                cell = ws.cell(row=2, column=col_idx)
                formula_original = None
                
                if hasattr(cell, 'value') and cell.value:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_original = cell.value
                
                if formula_original:
                    # Traduzir fórmula com suporte para referências entre abas
                    formula_traduzida = self._translate_formula_multi_sheet(formula_original, header_traducao)
                    
                    # Armazenar com prefixo da aba se necessário
                    formula_key = f"{aba_nome}.{header_resultado}" if len(abas_config) > 1 else header_resultado
                    
                    all_formulas[formula_key] = {
                        'formula_original': formula_original,
                        'formula_traduzida': formula_traduzida,
                        'aba': aba_nome,
                        'coluna_idx': col_idx,
                        'mapeamento': header_traducao,
                        'referencias_externas': self._extract_external_references(formula_original)
                    }
                    
                    self.logger.info(f"Fórmula extraída: {formula_key}")
        
        wb.close()
        
        # Analisar dependências entre fórmulas
        self._analyze_formula_dependencies(all_formulas)
        
        return all_formulas
    
    def _translate_formula_multi_sheet(self, formula: str, translation_map: Dict[str, str]) -> str:
        """
        Traduz fórmulas com suporte para referências entre abas
        Ex: Beneficios!B2 → VALE_REFEICAO
        """
        translated = formula
        
        # Padrão para referências com aba
        sheet_ref_pattern = re.compile(r'(\w+)!([A-Z]+)(\d+)')
        
        # Processar referências com aba primeiro
        for match in reversed(list(sheet_ref_pattern.finditer(formula))):
            sheet_name = match.group(1)
            col_letter = match.group(2)
            row_number = match.group(3)
            full_ref = f"{sheet_name}!{col_letter}"
            
            if full_ref in translation_map:
                new_name = translation_map[full_ref]
                start, end = match.span()
                translated = translated[:start] + new_name + translated[end:]
        
        # Depois processar referências locais (sem aba)
        local_refs = list(self.cell_ref_pattern.finditer(translated))
        for match in reversed(local_refs):
            col_letter = match.group(1)
            
            # Verificar se não faz parte de uma referência com aba
            pos = match.start()
            if pos > 0 and translated[pos-1] == '!':
                continue
                
            if col_letter in translation_map:
                new_name = translation_map[col_letter]
                start, end = match.span()
                translated = translated[:start] + new_name + translated[end:]
        
        return translated
    
    def _extract_external_references(self, formula: str) -> List[str]:
        """
        Extrai lista de referências externas (outras abas) de uma fórmula
        """
        sheet_ref_pattern = re.compile(r'(\w+)!')
        matches = sheet_ref_pattern.findall(formula)
        return list(set(matches))  # Remover duplicatas
    
    def _analyze_formula_dependencies(self, formulas: Dict[str, Dict[str, Any]]) -> None:
        """
        Analisa dependências entre fórmulas
        """
        for formula_key, formula_info in formulas.items():
            dependencies = []
            formula_text = formula_info['formula_traduzida']
            
            # Procurar por outras colunas referenciadas
            for other_key, other_info in formulas.items():
                if other_key != formula_key:
                    # Verificar se o nome da coluna aparece na fórmula
                    col_name = other_key.split('.')[-1] if '.' in other_key else other_key
                    if col_name in formula_text:
                        dependencies.append(other_key)
            
            formula_info['depends_on'] = dependencies
    
    def _find_column_by_header(self, worksheet, header_name: str) -> Optional[int]:
        """
        Encontra o índice da coluna pelo nome do header
        
        Args:
            worksheet: Planilha openpyxl
            header_name: Nome do header a procurar
            
        Returns:
            Índice da coluna (1-based) ou None se não encontrado
        """
        # Procurar na primeira linha
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value and str(cell_value).strip() == header_name:
                return col
        
        return None
    
    def _translate_formula(self, formula: str, translation_map: Dict[str, str]) -> str:
        """
        Traduz referências de célula em uma fórmula
        
        Args:
            formula: Fórmula original
            translation_map: Mapeamento de letras de coluna para nomes
            
        Returns:
            Fórmula traduzida
        """
        translated = formula
        
        # Encontrar todas as referências de célula
        matches = list(self.cell_ref_pattern.finditer(formula))
        
        # Processar de trás para frente para não bagunçar os índices
        for match in reversed(matches):
            col_letter = match.group(1)
            row_number = match.group(2)
            full_ref = match.group(0)
            
            # Verificar se temos tradução para esta coluna
            if col_letter in translation_map:
                new_name = translation_map[col_letter]
                # Substituir apenas a referência específica
                start, end = match.span()
                translated = translated[:start] + new_name + translated[end:]
            else:
                self.logger.debug(f"Sem tradução para coluna '{col_letter}'")
        
        return translated
    
    def extract_all_formulas_from_sheet(self, file_path: str, sheet_name: Optional[str] = None,
                                      row_start: int = 2, row_end: Optional[int] = None) -> pd.DataFrame:
        """
        Extrai todas as fórmulas de uma planilha (função auxiliar para debug/análise)
        
        Args:
            file_path: Caminho do arquivo Excel
            sheet_name: Nome da aba
            row_start: Linha inicial (padrão 2, assumindo headers na linha 1)
            row_end: Linha final (None para todas)
            
        Returns:
            DataFrame com informações sobre as fórmulas encontradas
        """
        wb = load_workbook(file_path, data_only=False)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        formulas_data = []
        
        # Determinar range de linhas
        max_row = row_end if row_end else ws.max_row
        
        # Iterar sobre células
        for row in range(row_start, min(max_row + 1, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Obter header da coluna
                    header = ws.cell(row=1, column=col).value
                    
                    formulas_data.append({
                        'linha': row,
                        'coluna': openpyxl.utils.get_column_letter(col),
                        'header': header,
                        'formula': cell.value,
                        'celula': f"{openpyxl.utils.get_column_letter(col)}{row}"
                    })
        
        wb.close()
        
        return pd.DataFrame(formulas_data)
    
    def validate_translation_map(self, formula: str, translation_map: Dict[str, str]) -> List[str]:
        """
        Valida se todas as referências em uma fórmula têm tradução
        
        Args:
            formula: Fórmula a validar
            translation_map: Mapeamento de tradução
            
        Returns:
            Lista de referências sem tradução
        """
        untranslated = []
        
        matches = self.cell_ref_pattern.finditer(formula)
        
        for match in matches:
            col_letter = match.group(1)
            if col_letter not in translation_map:
                untranslated.append(col_letter)
        
        return list(set(untranslated))  # Remover duplicatas
    
    def _extract_auto_discovery(self, config: Dict[str, Any], key_column: str) -> Dict[str, Dict[str, Any]]:
        """
        Extrai fórmulas usando descoberta automática
        """
        file_path = config['caminho']
        sample_row = config.get('linha_amostra', 2)
        lightweight = config.get('lightweight', False)
        columns_per_sheet = config.get('columns_per_sheet')
        max_cols = config.get('max_cols', 300 if not lightweight else 50)
        
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        mode = "[LIGHTWEIGHT] " if lightweight else ""
        self.logger.info(f"{mode}Usando descoberta automática de fórmulas em: {file_path}")
        
        # Usar o descobridor automático
        discoverer = FormulaAutoDiscovery()
        
        # Descobrir todas as fórmulas
        if lightweight:
            discovered_formulas = discoverer.discover_all_sheets_lightweight(
                file_path, 
                sample_row,
                columns_per_sheet=columns_per_sheet,
                max_cols_per_sheet=max_cols
            )
        else:
            discovered_formulas = discoverer.discover_all_sheets(file_path, sample_row)
        
        # Processar e traduzir as fórmulas descobertas
        all_formulas = {}
        
        for formula_key, formula_info in discovered_formulas.items():
            # Usar o mapeamento descoberto automaticamente
            mapeamento = formula_info.get('header_traducao', {})
            
            # Se não houver mapeamento, criar um baseado nas dependências
            if not mapeamento:
                mapeamento = self._create_auto_mapping(formula_info)
            
            # Traduzir a fórmula
            formula_traduzida = self._translate_formula_multi_sheet(
                formula_info['formula_original'], 
                mapeamento
            )
            
            all_formulas[formula_key] = {
                'formula_original': formula_info['formula_original'],
                'formula_traduzida': formula_traduzida,
                'aba': formula_info['aba'],
                'coluna': formula_info['column'],
                'coluna_idx': openpyxl.utils.column_index_from_string(formula_info['column']),
                'mapeamento': mapeamento,
                'referencias_externas': formula_info.get('sheet_references', []),
                'depends_on': self._extract_dependencies_from_translated(formula_traduzida),
                'auto_discovered': True
            }
            
            self.logger.info(f"Fórmula descoberta e traduzida: {formula_key}")
        
        # Analisar dependências entre fórmulas
        self._analyze_formula_dependencies(all_formulas)
        
        return all_formulas
    
    def _create_auto_mapping(self, formula_info: Dict[str, Any]) -> Dict[str, str]:
        """
        Cria mapeamento automático baseado nas dependências descobertas
        """
        mapeamento = {}
        formula = formula_info['formula_original']
        
        # Para cada dependência, criar entrada no mapeamento
        for dep in formula_info.get('dependencies', []):
            if '!' in dep:
                # Referência com aba (Sheet!Column)
                sheet, col_or_name = dep.split('!', 1)
                # Encontrar a letra da coluna na fórmula
                pattern = rf'{re.escape(sheet)}!([A-Z]+)\d+'
                matches = re.findall(pattern, formula)
                for col_letter in set(matches):
                    mapeamento[f"{sheet}!{col_letter}"] = col_or_name
            else:
                # Referência local
                # Encontrar letras de coluna na fórmula
                pattern = r'(?<![A-Z!])([A-Z]+)\d+'
                matches = re.findall(pattern, formula)
                for col_letter in set(matches):
                    if col_letter not in mapeamento:
                        mapeamento[col_letter] = dep
        
        return mapeamento
    
    def _translate_formula_auto(self, formula: str, mapeamento: Dict[str, str], 
                               dependencies: List[str]) -> str:
        """
        Traduz fórmula com mapeamento automático
        """
        translated = formula
        
        # Ordenar chaves por comprimento (maiores primeiro) para evitar substituições parciais
        sorted_keys = sorted(mapeamento.keys(), key=len, reverse=True)
        
        for key in sorted_keys:
            value = mapeamento[key]
            # Padrão para encontrar a referência completa
            if '!' in key:
                # Referência com aba
                pattern = rf'{re.escape(key)}(\d+)'
            else:
                # Referência local
                pattern = rf'(?<![A-Z!]){re.escape(key)}(\d+)'
            
            # Substituir mantendo o número da linha
            translated = re.sub(pattern, value, translated)
        
        return translated
    
    def _extract_dependencies_from_translated(self, formula_traduzida: str) -> List[str]:
        """
        Extrai lista de dependências de uma fórmula traduzida
        """
        # Padrão para nomes de colunas (palavras em maiúsculas com _)
        pattern = r'\b[A-Z][A-Z_]*[A-Z]\b'
        dependencies = re.findall(pattern, formula_traduzida)
        
        # Remover funções Excel conhecidas
        excel_functions = {'IF', 'SUM', 'SUMIF', 'VLOOKUP', 'AVERAGEIF', 'AND', 'OR', 'NOT'}
        dependencies = [d for d in dependencies if d not in excel_functions]
        
        return list(set(dependencies))  # Remover duplicatas